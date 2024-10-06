from openpyxl import Workbook, load_workbook
import openpyxl

wb = openpyxl.load_workbook('Database.xlsx')
'''
wb = openpyxl.load_workbook('Database.xlsx')
orders_sheet = wb['Orders']

orders_sheet = wb.create_sheet(title='Orders')
orders_sheet.append(['Customer', 'Vendor', 'Value', 'Status', 'Description', 'Quantity', 'OrderID'])

customers_sheet = wb.create_sheet(title='Customers')
customers_sheet.append(['Name', 'Phone Number', 'Referral', 'Orders Placed',
                         'Total Spending', 'Additional Information', 'Past Orders'])

vendors_sheet = wb.create_sheet(title="Vendors")
vendors_sheet = wb['Vendors']
vendors_sheet.append(['Name', 'Phone Number', 'Address', 'Status', 'Additional Information',
                      'Past Orders', 'Current Orders'])

wb.save('Database.xlsx')
'''
def add_entry(active_sheet, entry):
    active_sheet.append(entry)
    print(entry + "Added to database")
    wb.save('Database.xlsx')

def delete_entry(active_sheet, entry_row):
    active_sheet.delete_rows(entry_row)
    wb.save('Database.xlsx')

def find_entry_info(sheet, entry, entry_col):
    active_sheet = wb[f'{sheet}']
    for row in active_sheet.iter_rows(min_row=1, max_row=active_sheet.max_row,
                                       min_col=entry_col, max_col=entry_col):
        for cell in row:
            if cell.value == entry:
                #print(f"Found '{entry}' at row {cell.row}, column {cell.column}")
                return cell.row

def edit_entry(active_sheet, entry, entry_col, entry_row):
    active_sheet.cell(row=entry_row, column=entry_col, value=entry)
    wb.save('Database.xlsx')

def show_entry(active_sheet, entry_row):
    
    data_row = []
    for column in active_sheet:
        value = active_sheet.cell(row=entry_row, column=column).value
        data_row.append(value)

    return data_row

def id_from_row(row_num):
    wb = openpyxl.load_workbook('Database.xlsx')
    orders_sheet = wb['Orders']

    orderid = orders_sheet.cell(row=row_num, column=6).value

    return orderid