import openpyxl

class Customer:
    def __init__(self, Name=None, Phone_Num=None, Referral=None, AdditionalInfo=None):
        self.Name = Name
        self.PhoneNumber = Phone_Num
        self.Referral = Referral
        self.OrdersPlaced = 0.0
        self.TotalSpending = 0.0
        self.AdditionalInfo = AdditionalInfo
        self.PastOrders = []

    def get_name(self):
        return self.Name

    def get_phone_number(self):
        return self.PhoneNumber

    def get_referral(self):
        return self.Referral

    def get_orders_placed(self):
        return self.OrdersPlaced

    def get_total_spending(self):
        return self.TotalSpending

    def get_additional_info(self):
        return self.AdditionalInfo

    def get_past_orders(self):
        return self.PastOrders

    def set_name(self, name):
        self.Name = name

    def set_phone_number(self, phone_number):
        self.PhoneNumber = phone_number

    def set_referral(self, referral):
        self.Referral = referral

    def set_orders_placed(self, orders_placed):
        self.OrdersPlaced = orders_placed

    def set_additional_info(self, additional_info):
        self.AdditionalInfo = additional_info

    def add_past_order(self, past_order):
        self.PastOrders.append(past_order)

    def store_customer(self):
    
        wb = openpyxl.load_workbook('Database.xlsx')
        ws = wb['Customers']

        Past_Orders = ','.join(self.PastOrders)

        ws.append([self.Name, self.PhoneNumber, self.Referral, self.OrdersPlaced, self.TotalSpending, self.AdditionalInfo, Past_Orders])

        wb.save('Database.xlsx')

    def calc_orders_placed(self, past_orders):
        return len(past_orders)

    def calc_total_spending(self, past_orders):

        #Access Excel Database with Order Information
        wb = openpyxl.load_workbook('Database.xlsx')
        order_sheet = wb['Orders']

        total_spending = 0

        for order in past_orders:
            #Resets quantities after every order
            order_value = 0
            order_quantity = 0
            order_spending = 0
            for row in order_sheet.iter_rows(min_row=2, max_row=order_sheet.max_row):
                #Defines orderid
                orderid = row[6].value
                if order == orderid:
                    #Assigns values to quantities based on order
                    order_value = float(row[2].value)
                    order_quantity = float(row[5].value)
                    order_spending = order_value * order_quantity
                    total_spending = total_spending + order_spending
                else:
                    continue

        wb.save('Database.xlsx')
        return total_spending
