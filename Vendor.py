import openpyxl

class Vendor:
    def __init__(self, Name=None, PhoneNumber=None, Address=None, AdditionalInfo=None):
        self.Name = Name
        self.PhoneNumber = PhoneNumber
        self.Address = Address
        self.Status = None
        self.AdditionalInfo = AdditionalInfo
        self.past_orders = []
        self.current_orders = []

    def get_name(self):
        return self.Name

    def get_phone_number(self):
        return self.PhoneNumber

    def get_address(self):
        return self.Address

    def get_status(self):
        return self.Status

    def get_additional_info(self):
        return self.AdditionalInfo

    def set_name(self, name):
        self.Name = name

    def set_phone_number(self, phone_number):
        self.PhoneNumber = phone_number

    def set_address(self, address):
        self.Address = address

    def set_additional_info(self, additional_info):
        self.AdditionalInfo = additional_info

    def store_vendor(self):
    
        wb = openpyxl.load_workbook('Database.xlsx')
        ws = wb['Vendors']

        Past_Orders = ','.join(self.past_orders)
        Current_Orders = ','.join(self.current_orders)

        ws.append([self.Name, self.PhoneNumber, self.Address, self.Status, self.AdditionalInfo, Past_Orders, Current_Orders])

        wb.save('Database.xlsx')

    def add_current_order(vendor, order, current_orders):
            
        #Access Vendors DB
        wb = openpyxl.load_workbook('Database.xlsx')
        vendor_sheet = wb['Vendors']

        current_orders.append(order)

        for row in vendor_sheet:
            if row[0].value == vendor:
                row[6] = current_orders
            else:
                continue
        wb.save('Database.xlsx')

    #Checks whether an order belongs to current orders or past orders based on order status
    def manage_order_storage(current_orders, past_orders):
            
        #Load workbook and worksheet
        wb = openpyxl.load_workbook('Database.xlsx')
        order_sheet = wb['Orders']
        vendor_sheet = wb['Vendors']
            
        for order in current_orders:
            for row in order_sheet:
                current_orderid = row[6].value
                if order == current_orderid:
                    status = row[3].value
                    #Check to see if order is completed (status would be 100 - representing 100%)
                    if status >= 100:
                        past_orders.append(order)
                        current_orders.remove(order)
                    else:
                        continue
                else:
                    continue

        wb.save('Database.xlsx')
        return current_orders, past_orders
    

    def set_status():
        pass
'''
        def add_past_order(self, past_order):
            self.PastOrders.append(past_order)

        def calc_orders_placed(self):
            return len(self.PastOrders)

        def calc_total_spending(self):
            return sum(self.PastOrders)

'''