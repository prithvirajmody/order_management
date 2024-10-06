from openpyxl import Workbook, load_workbook
import random
import string

import openpyxl

class Order:
    def __init__(self, Client=None, Vendor=None, Value=0.0, Status=None, Description=None,
                  Quantity=0, due_date=None):
        #Stores Customer Phone Number
        self.Client = Client
        #Stores Vendor Phone Number
        self.Vendor = Vendor
        #Stores Rupee Value of Order
        self.Value = Value
        #Stores Completion Status of Order
        self.Status = Status
        #Date stored as a string in format "ddmmyyyy"
        self.due_date = due_date
        self.Description = Description
        self.Quantity = Quantity
        #Not entered by user, generated automatically by program and assigned to this variable
        self.OrderID = None

    def get_client(self):
        return self.Client
    
    def set_client(self, client):
        self.Client = client

    def get_vendor(self):
        return self.Vendor

    def get_value(self):
        return self.Value

    def get_status(self):
        return self.Status

    def get_description(self):
        return self.Description

    def get_quantity(self):
        return self.Quantity

    def set_vendor(self, vendor):
        self.Vendor = vendor

    def set_value(self, value):
        self.Value = value

    def set_status(self, status):
        self.Status = status

    def set_description(self, description):
        self.Description = description

    def set_quantity(self, quantity):
        self.Quantity = quantity

    def link_client(self, Client, OrderID):

        #Load workbook and worksheet
        wb = openpyxl.load_workbook('Database.xlsx')
        customer_sheet = wb['Customers']

        #Access Phone Number Column in database
        for row in customer_sheet.iter_rows(min_row=2, max_row=customer_sheet.max_row):
            #Select the value from the second column in the row (where phone number is stored)
            customer = row[1].value
            if customer == Client:
                #Check if there are already past orders stored in the row
                if row[6].value is not None:
                    current_orders = row[6].value
                    new_orders = f"{current_orders}, {OrderID}"
                else:
                    new_orders = f"{OrderID}"

                customer_sheet.cell(row=row[0].row, column=7, value=new_orders)
                wb.save('Database.xlsx')

            else:
                continue



    def link_vendor(self, Vendor, OrderID):

        #Load workbook and worksheet
        wb = openpyxl.load_workbook('Database.xlsx')
        vendor_sheet = wb['Vendors']

        #Access Phone Number Column in database
        for row in vendor_sheet.iter_rows(min_row=2, max_row=vendor_sheet.max_row):
            #Select the value from the second column in the row
            vendor_id = row[1].value
            if vendor_id == Vendor:
                #Check if there are already current orders stored in the row
                if row[6].value is not None:
                    current_orders = row[6].value
                    new_orders = f"{current_orders}, {OrderID}"
                else:
                    new_orders = f"{OrderID}"

                vendor_sheet.cell(row=row[0].row, column=7, value=new_orders)
                wb.save('Database.xlsx')

            else:
                continue

    def status_notify(self):
        pass

    def create_order_id(self):
    # Create Random OrderID
        characters = string.ascii_letters + string.digits
        OrderID = ''.join(random.choice(characters) for _ in range(8))

        # Check if OrderID already used
        wb = openpyxl.load_workbook('Database.xlsx')
        ws = wb['Orders']

        # Collect existing order IDs in a set
        existing_order_ids = {cell.value for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=6, max_col=6) for cell in row}

        if OrderID in existing_order_ids:
            # If OrderID already exists, try again
            self.create_order_id()
        else:
            # Assign unique OrderID to Order
            self.OrderID = OrderID

        wb.save('Database.xlsx')
        return OrderID


    #Opens delete confirmation page
    def show_delete_page(self):
        pass

    #Called to store order in database (When submit button pressed in UI)
    def storeOrder(self):
    
        wb = openpyxl.load_workbook('Database.xlsx')
        ws = wb['Orders']

        ws.append([self.Client, self.Vendor, self.Value, self.Status, self.Description, self.Quantity, self.OrderID, self.due_date])

        wb.save('Database.xlsx')

#order = Order("0987654321", "1234567890", 7500, 10, "test2", 5, "22022024")
#orderid = order.create_order_id()
#print(orderid)
#order.link_vendor("1234567890", orderid)