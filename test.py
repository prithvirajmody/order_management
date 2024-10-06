import openpyxl

def calc_total_spending(past_orders):

    #Access Excel Database with Order Information
    wb = openpyxl.load_workbook('Database.xlsx')
    order_sheet = wb['Orders']

    total_spending = 0

    for order in past_orders:
        #Resets quantities after every order
        order_value = 0
        order_quantity = 0
        order_spending = 0
        for row in order_sheet.iter_rows(min_row=2, max_row=order_sheet.max_row):
            #Defines orderid
            orderid = row[6].value
            if order == orderid:
                #Assigns values to quantities based on order
                order_value = float(row[2].value)
                order_quantity = float(row[5].value)
                order_spending = order_value * order_quantity
                total_spending = total_spending + order_spending
            else:
                continue

    wb.save('Database.xlsx')
    return total_spending


past_orders = ['fjKdCEim', 'n4QY1vqV', '14jUUOox', 'aGuMo4yk']
ts = calc_total_spending(past_orders=past_orders)
print(ts)