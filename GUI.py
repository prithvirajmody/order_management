#External Libraries Imported
import PySimpleGUI as sg
import openpyxl
from datetime import datetime, timedelta

#Libraries created by me for the project
from Order import Order
from Customer import Customer
from Vendor import Vendor 
from data_sort import sorting_functions
from Exceldb import find_entry_info, delete_entry, add_entry
from Queue import Queue

sg.theme('GreenMono')

def login(username, password, window):
    
    if username == 'SmritiMody' and password == 'Awesome2307':
        window.close()
        homepage_window()
    else:
        sg.popup_no_titlebar("Incorrect Credentials")

def main_window():

    layout = [
        [sg.Text("Login", pad=15, expand_x=True, font=('Arial Bold', 22), justification='center')],
        [sg.Text("Username:", pad=10, expand_x=True, font=('Times New Roman', 15), justification='center')],
        [sg.InputText(key="-USRNME-", pad=10)],
        [sg.Text("Password:", pad=10, expand_x=True, font=('Times New Roman', 15), justification='center')],
        [sg.InputText(key="-PSSWD-", pad=10, password_char='*')],
        [sg.Button("Login", pad=15, button_color='#041c11', mouseover_colors='#02b8fa', size=(15, 1))],
        [sg.Exit(button_color='#041c11', size=(15, 1))]
    ]

    window = sg.Window("Login Page", layout, size=(400, 350), element_justification='c')

    while True:
        event, values = window.read()
        if event in (sg.WINDOW_CLOSED, "Exit"):
            break
        if event == "Login":
            login(username=values["-USRNME-"], password=values["-PSSWD-"], window=window)

    window.close()

def homepage_window():
    
    homepage_layout = [
        [sg.Exit(button_color='#041c11', size=(5, 3)), sg.Text("Password Publications", pad=15, expand_x=True, font=('Arial Bold', 22), justification='center')],
        [sg.Button("Orders", pad=0, button_color='#041c11', mouseover_colors='#02b8fa', size=(15, 3), font=('Times New Roman', 15))],
        [sg.Button("Customers", pad=10, button_color='#041c11', mouseover_colors='#02b8fa', size=(15, 3), font=('Times New Roman', 15))],
        [sg.Button("Vendors", pad=0, button_color='#041c11', mouseover_colors='#02b8fa', size=(15, 3), font=('Times New Roman', 15))]
    ]

    window = sg.Window("Home Page", homepage_layout, size=(450, 350), element_justification='c')

    while True:
        event, values = window.read()
        if event in (sg.WINDOW_CLOSED, "Exit"):
            break
        if event == "Orders":
            window.close()
            orders_window(queue=False)
        if event == "Customers":
            window.close()
            customers_window(queue=False)
        if event == "Vendors":
            window.close()
            vendors_window(queue=False)
    window.close()

def orders_window(queue):

    wb = openpyxl.load_workbook('Database.xlsx')
    ws = wb['Orders']

    # Get the number of rows in the Excel sheet
    num_rows = ws.max_row

    # Create lists to hold Text and Button elements dynamically
    text_elements = []
    button_elements = []

    if queue == False:
        for row_num in range(2, num_rows + 1):
            order_text = sg.Text(f"Order {row_num - 1}", pad=5)
            button = sg.Button("View/Edit Order", key=f"BUTTON{row_num - 1}")
            text_elements.append(order_text)
            button_elements.append(button)
    
    if queue == True:
        restored = Queue.deserialize('queue_state.pkl')

        lst = []

        while restored.is_empty() == False:
            temp = restored.dequeue()
            lst.append(temp)

        for i in range(0, len(lst)):
            order_text = sg.Text(f"Order {lst[i]}", pad=5)
            button = sg.Button("View/Edit Order", key=f"BUTTON{lst[i] - 1}")
            text_elements.append(order_text)
            button_elements.append(button)

    # Create the column layout dynamically
    col_layout = list(zip(text_elements, button_elements))

    layout = [
        [sg.Button("Back To Menu", button_color='#041c11', size=(7, 3), mouseover_colors='#02b8fa'), sg.Text("Orders", pad=10, expand_x=True, font=('Arial Bold', 24), justification='center'), sg.Exit(button_color='#041c11', size=(7, 3))],
        [sg.Button("Add New", pad=10, button_color='#041c11', mouseover_colors='#02b8fa', size=(15, 1)), sg.Button("Sort Entries", pad=10, button_color='#041c11', mouseover_colors='#02b8fa', size=(15, 1))],
        [sg.Column(col_layout, scrollable=True, size=(500, 275), vertical_scroll_only=True, element_justification='center')]
    ]

    window = sg.Window("Orders Page", layout, size=(400, 400), element_justification='c')

    while True:
        event, values = window.read()
        if event in (sg.WINDOW_CLOSED, "Exit"):
            break
        if event == "Sort Entries":
            window.close()
            sort_orders_window()
        if event == "Add New":
            window.close()
            create_order_window()
        if event.startswith("BUTTON"):
            # Extract the row_num from the event key
            row_num = int(event.replace("BUTTON", ""))
            window.close()
            view_order_window(row_num + 1)
        if event == "Back To Menu":
            window.close()
            homepage_window()

    window.close()

def customers_window(queue):

    wb = openpyxl.load_workbook('Database.xlsx')
    ws = wb['Customers']

    # Get the number of rows in the Excel sheet
    num_rows = ws.max_row

    # Create lists to hold Text and Button elements dynamically
    text_elements = []
    button_elements = []

    if queue == False:
        for row_num in range(2, num_rows + 1):
            order_text = sg.Text(f"Customer {row_num - 1}", pad=5)
            button = sg.Button("View/Edit Customer", key=f"BUTTON{row_num - 1}")
            text_elements.append(order_text)
            button_elements.append(button)

    if queue == True:
        restored = Queue.deserialize('queue_state.pkl')

        lst = []

        while restored.is_empty() == False:
            temp = restored.dequeue()
            lst.append(temp)

        for i in range(0, len(lst)):
            order_text = sg.Text(f"Customer {lst[i] - 1}", pad=5)
            button = sg.Button("View/Edit Customer", key=f"BUTTON{lst[i] - 1}")
            text_elements.append(order_text)
            button_elements.append(button)

    # Create the column layout dynamically
    col_layout = list(zip(text_elements, button_elements))

    layout = [
        [sg.Button("Back To Menu", button_color='#041c11', size=(7, 3), mouseover_colors='#02b8fa'), sg.Text("Customers", pad=10, expand_x=True, font=('Arial Bold', 24), justification='center'), sg.Exit(button_color='#041c11', size=(7, 3))],
        [sg.Button("Add New", pad=10, button_color='#041c11', mouseover_colors='#02b8fa', size=(15, 1)), sg.Button("Sort", pad=10, button_color='#041c11', mouseover_colors='#02b8fa', size=(15, 1))],
        [sg.Column(col_layout, scrollable=True, size=(500, 275), vertical_scroll_only=True, element_justification='center')]
    ]

    window = sg.Window("Customers Page", layout, size=(400, 400), element_justification='c')

    while True:
        event, values = window.read()
        if event in (sg.WINDOW_CLOSED, "Exit"):
            break
        if event == "Sort":
            window.close()
            sort_customers_window()
        if event == "Add New":
            window.close()
            add_customer_window()
        if event.startswith("BUTTON"):
            row_num = int(event.replace("BUTTON", ""))
            window.close()
            view_customer_window(row_num + 1)
        if event == "Back To Menu":
            window.close()
            homepage_window()

    window.close()

def vendors_window(queue):

    wb = openpyxl.load_workbook('Database.xlsx')
    ws = wb['Vendors']

    # Get the number of rows in the Excel sheet
    num_rows = ws.max_row

    # Create lists to hold Text and Button elements dynamically
    text_elements = []
    button_elements = []

    if queue == False:
        for row_num in range(2, num_rows + 1):
            order_text = sg.Text(f"Vendor {row_num - 1}", pad=5)
            button = sg.Button("View/Edit Vendor", key=f"BUTTON{row_num - 1}")
            text_elements.append(order_text)
            button_elements.append(button)

    if queue == True:
        restored = Queue.deserialize('queue_state.pkl')

        lst = []

        while restored.is_empty() == False:
            temp = restored.dequeue()
            lst.append(temp)

        for i in range(0, len(lst)):
            order_text = sg.Text(f"Vendor {lst[i] - 1}", pad=5)
            button = sg.Button("View/Edit Vendor", key=f"BUTTON{lst[i] - 1}")
            text_elements.append(order_text)
            button_elements.append(button)

    # Create the column layout dynamically
    col_layout = list(zip(text_elements, button_elements))

    layout = [
        [sg.Button("Back To Menu", button_color='#041c11', size=(7, 3), mouseover_colors='#02b8fa'),
          sg.Text("Vendors", pad=10, expand_x=True, font=('Arial Bold', 24), justification='center'),
            sg.Exit(button_color='#041c11', size=(7, 3))],
        [sg.Button("Add New", pad=10, button_color='#041c11', mouseover_colors='#02b8fa', size=(15, 1)), 
         sg.Button("Sort", pad=10, button_color='#041c11', mouseover_colors='#02b8fa', size=(15, 1))],
        [sg.Column(col_layout, scrollable=True, size=(500, 275), vertical_scroll_only=True,
                    element_justification='center')]
    ]

    window = sg.Window("Vendors Page", layout, size=(400, 350), element_justification='c')

    while True:
        event, values = window.read()
        if event in (sg.WINDOW_CLOSED, "Exit"):
            break
        if event == "Sort":
            window.close()
            sort_vendors_window()
        if event == "Add New":
            window.close()
            add_vendor_window()
        if event.startswith("BUTTON"):
            row_num = int(event.replace("BUTTON", ""))
            window.close()
            view_vendor_window(row_num + 1)
        if event == "Back To Menu":
            window.close()
            homepage_window()

    window.close()

def sort_orders_window():
    button_states = {
        "Value of Order": 0,
        "Order Status": 0,
        "Due Date": 0
    }

    layout = [
        [sg.Text("Sort By",  pad=15, expand_x=True, font=('Arial Bold', 24), justification='center')],
        [sg.Button("Value of Order", pad=0, button_color=('#041c11', '#02b8fa'), mouseover_colors=('#02b8fa', '#041c11'), size=(15, 2), font=('Times New Roman', 15), key="Value of Order")],
        [sg.Button("Order Status", pad=15, button_color=('#041c11', '#02b8fa'), mouseover_colors=('#02b8fa', '#041c11'), size=(15, 2), font=('Times New Roman', 15), key="Order Status")],
        [sg.Button("Due Date", pad=0, button_color=('#041c11', '#02b8fa'), mouseover_colors=('#02b8fa', '#041c11'), size=(15, 2), font=('Times New Roman', 15), key="Due Date")],
        [sg.Text("Least/Earliest First", font=('Times New Roman', 12)), sg.Slider(range=(0, 1), orientation='h', default_value=0, resolution=1, disable_number_display=True, pad=10, key='-SLIDER-', size=(18, 10), background_color='#041c11'), sg.Text("Most/Latest First", font=('Times New Roman', 12))],
        [sg.Button("Back", pad=30, button_color='#041c11', mouseover_colors='#02b8fa', size=(15, 1)), sg.Button("Confirm", pad=30, button_color='#041c11', mouseover_colors='#02b8fa', size=(15, 1))],
        [sg.Exit(button_color='#041c11', size=(15, 1))]
    ]

    window = sg.Window("Sort Orders Page", layout, size=(500, 450), element_justification='c')

    while True:
        event, values = window.read()

        if event in (sg.WINDOW_CLOSED, "Exit"):
            break

        if event == "Back":
            window.close()
            orders_window(queue=False)

        # Toggle button color
        if event in button_states:
            button_name = event

            # Reset all buttons to 0 except the one clicked
            for key in button_states:
                if key != button_name:
                    button_states[key] = 0

            # Toggle the selected button
            button_states[button_name] = 1

            # Update button colors based on states
            for key, state in button_states.items():
                new_color = ('#02b8fa', '#041c11') if state == 1 else ('#041c11', '#02b8fa')
                window[key].update(button_color=new_color)

        if event == "Confirm":
            selected_criteria = [key for key, value in button_states.items() if value == 1][0]
            sorting_order = "ascending" if values['-SLIDER-'] == 0 else "descending"

            sort = sorting_functions()

            wb = openpyxl.load_workbook('Database.xlsx')
            ws = wb['Orders']

            data = []

            if selected_criteria == "Value of Order":

                for row in ws.iter_rows(min_row=2, values_only=True):
                    row_value = row[2]

                    try:
                        float_value = float(row_value)
                    except ValueError:
                        float_value = 0.0  # Assign a default value of 0.0 for empty strings

                    if row_value is not None:
                        data.append(float_value)
                    else:
                        break

                if sorting_order == "ascending":

                    sorted_data = sort.number_sort(data, True)

                    data_order = []

                    for i in range(0, len(sorted_data)):
                        temp = sorted_data[i]

                        active_sheet = wb['Orders']

                        for i in active_sheet.iter_rows(min_row=2, max_row=active_sheet.max_row, min_col=3, max_col=3):
                            for cell in i:
                                if cell.value == temp and cell.row not in data_order:
                                    row = cell.row
                                    data_order.append(row)

                    queue = Queue()

                    for i in range(0, len(data_order)):
                        queue.enqueue(data_order[i])

                    queue.serialize('queue_state.pkl')
                    
                    window.close()
                    orders_window(queue=True)

                if sorting_order == "descending":

                    sorted_data = sort.number_sort(data, False)

                    data_order = []

                    for i in range(0, len(sorted_data)):
                        temp = sorted_data[i]
                        temp = float(temp)

                        active_sheet = wb['Orders']

                        for row in active_sheet.iter_rows(min_row=2, max_row=active_sheet.max_row, min_col=3, max_col=3):
                            for cell in row:
                                if cell.value == temp and cell.row not in data_order:
                                    row_val = cell.row
                                    data_order.append(row_val)

                    queue = Queue()

                    for i in range(0, len(data_order)):
                        queue.enqueue(data_order[i])

                    queue.serialize('queue_state.pkl')

                    window.close()
                    orders_window(queue=True)

            if selected_criteria == "Order Status":

                for row in ws.iter_rows(values_only=True):
                    row_value = row[3]

                    try:
                        float_value = float(row_value)
                    except ValueError:
                        float_value = 0.0  # Assign a default value of 0.0 for empty strings

                    if row_value is not None:
                        data.append(float_value)
                    else:
                        break                

                if sorting_order == "ascending":

                    sorted_data = sort.number_sort(data, True)

                    data_order = []

                    for i in range(0, len(sorted_data)):
                        temp = sorted_data[i]

                        active_sheet = wb['Orders']

                        for row in active_sheet.iter_rows(min_row=2, max_row=active_sheet.max_row, min_col=4, max_col=4):
                            for cell in row:
                                if cell.value == temp and cell.row not in data_order:
                                    row_val = cell.row
                                    data_order.append(row_val)

                    queue = Queue()

                    for i in range(0, len(data_order)):
                        queue.enqueue(data_order[i])

                    queue.serialize('queue_state.pkl')

                    window.close()
                    orders_window(queue=True)

                if sorting_order == "descending":

                    sorted_data = sort.number_sort(data, False)

                    data_order = []

                    for i in range(0, len(sorted_data)):
                        temp = sorted_data[i]
                        temp = float(temp)
                        
                        active_sheet = wb['Orders']

                        for row in active_sheet.iter_rows(min_row=2, max_row=active_sheet.max_row, min_col=4, max_col=4):
                            for cell in row:
                                if cell.value == temp and cell.row not in data_order:
                                    row_val = cell.row
                                    data_order.append(row_val)

                    queue = Queue()

                    for i in range(0, len(data_order)):
                        queue.enqueue(data_order[i])

                    queue.serialize('queue_state.pkl')

                    window.close()
                    orders_window(queue=True)

            if selected_criteria == "Due Date":

                for row in ws.iter_rows(min_row=2, values_only=True):
                    row_value = row[7]

                    if row_value is not None and not isinstance(row_value, str):
                        row_value = str(row_value)
                    data.append(row_value)

                if sorting_order == "ascending":

                    sorted_data = sort.date_sort(data, True)

                    data_order = []

                    for i in range(0, len(sorted_data)):
                        temp = sorted_data[i]
                        temp = int(temp)
                        
                        active_sheet = wb['Orders']

                        for i in active_sheet.iter_rows(min_row=2, max_row=active_sheet.max_row, min_col=8, max_col=8):
                            for cell in i:
                                if cell.value == temp and cell.row not in data_order:
                                    row = cell.row
                                    data_order.append(row)
        
                    queue = Queue()

                    for i in range(0, len(data_order)):
                        queue.enqueue(data_order[i])

                    queue.serialize('queue_state.pkl')

                    window.close()
                    orders_window(queue=True)

                if sorting_order == "descending":

                    sorted_data = sort.date_sort(data, False)

                    data_order = []

                    for i in range(0, len(sorted_data)):
                        temp = sorted_data[i]
                        temp = int(temp)
                        
                        active_sheet = wb['Orders']

                        for i in active_sheet.iter_rows(min_row=2, max_row=active_sheet.max_row, min_col=8, max_col=8):
                            for cell in i:
                                if cell.value == temp and cell.row not in data_order:
                                    row = cell.row
                                    data_order.append(row)

                    queue = Queue()

                    for i in range(0, len(data_order)):
                        queue.enqueue(data_order[i])

                    queue.serialize('queue_state.pkl')

                    window.close()
                    orders_window(queue=True)

    window.close()

def sort_customers_window():
    button_states = {
        "Number of Orders": 0,
        "Name": 0,
        "Total Spending": 0
    }

    layout = [
        [sg.Text("Sort By",  pad=15, expand_x=True, font=('Arial Bold', 24), justification='center')],
        [sg.Button("Number of Orders", pad=0, button_color=('#041c11', '#02b8fa'), mouseover_colors=('#02b8fa', '#041c11'), size=(15, 2), font=('Times New Roman', 15), key="Number of Orders")],
        [sg.Button("Name", pad=15, button_color=('#041c11', '#02b8fa'), mouseover_colors=('#02b8fa', '#041c11'), size=(15, 2), font=('Times New Roman', 15), key="Name")],
        [sg.Button("Total Spending", pad=0, button_color=('#041c11', '#02b8fa'), mouseover_colors=('#02b8fa', '#041c11'), size=(15, 2), font=('Times New Roman', 15), key="Total Spending")],
        [sg.Text("Least/A First", font=('Times New Roman', 12)), sg.Slider(range=(0, 1), orientation='h', default_value=0, resolution=1, disable_number_display=True, pad=10, key='-SLIDER-', size=(18, 10), background_color='#041c11'), sg.Text("Most/Z First", font=('Times New Roman', 12))],
        [sg.Button("Back", pad=30, button_color='#041c11', mouseover_colors='#02b8fa', size=(15, 1)), sg.Button("Confirm", pad=30, button_color='#041c11', mouseover_colors='#02b8fa', size=(15, 1))],
        [sg.Exit(button_color='#041c11', size=(15, 1))]
    ]

    window = sg.Window("Sort Customers Page", layout, size=(500, 450), element_justification='c')

    while True:
        event, values = window.read()

        if event in (sg.WINDOW_CLOSED, "Exit"):
            break

        if event == "Back":
            window.close()
            customers_window(False)

        # Toggle button color
        if event in button_states:
            button_name = event

            # Reset all buttons to 0 except the one clicked
            for key in button_states:
                if key != button_name:
                    button_states[key] = 0

            # Toggle the selected button
            button_states[button_name] = 1

            # Update button colors based on states
            for key, state in button_states.items():
                new_color = ('#02b8fa', '#041c11') if state == 1 else ('#041c11', '#02b8fa')
                window[key].update(button_color=new_color)

        if event == "Confirm":
            selected_criteria = [key for key, value in button_states.items() if value == 1][0]
            sorting_order = "ascending" if values['-SLIDER-'] == 0 else "descending"

            sort = sorting_functions()

            wb = openpyxl.load_workbook('Database.xlsx')
            ws = wb['Customers']

            data = []

            if selected_criteria == "Number of Orders":

                for row in ws.iter_rows(min_row=2, values_only=True):
                    row_value = row[3]

                    try:
                        int_value = int(row_value)
                    except ValueError:
                        int_value = 0  # Assign a default value of 0 for empty strings

                    if row_value is not None:
                        data.append(int_value)
                    else:
                        break

                if sorting_order == "ascending":

                    sorted_data = sort.number_sort(data, True)
                    
                    data_order = []

                    for i in range(0, len(sorted_data)):
                        temp = sorted_data[i]
                        temp = int(temp)

                        active_sheet = wb['Customers']

                        for i in active_sheet.iter_rows(min_row=2, max_row=active_sheet.max_row, min_col=4, max_col=4):
                            for cell in i:
                                if cell.value == temp and cell.row not in data_order:
                                    row = cell.row
                                    data_order.append(row)

                    queue = Queue()

                    for i in range(0, len(data_order)):
                        queue.enqueue(data_order[i])

                    queue.serialize('queue_state.pkl')

                    window.close()
                    customers_window(queue=True)

                if sorting_order == "descending":

                    sorted_data = sort.number_sort(data, False)
                    
                    data_order = []

                    for i in range(0, len(sorted_data)):
                        temp = sorted_data[i]
                        temp = int(temp)

                        active_sheet = wb['Customers']

                        for i in active_sheet.iter_rows(min_row=2, max_row=active_sheet.max_row, min_col=4, max_col=4):
                            for cell in i:
                                if cell.value == temp and cell.row not in data_order:
                                    row = cell.row
                                    data_order.append(row)

                    queue = Queue()

                    for i in range(0, len(data_order)):
                        queue.enqueue(data_order[i])

                    queue.serialize('queue_state.pkl')

                    window.close()
                    customers_window(queue=True)

            if selected_criteria == "Total Spending":

                for row in ws.iter_rows(min_row=2, values_only=True):
                    row_value = row[4]

                    try:
                        float_value = float(row_value)
                    except ValueError:
                        float_value = 0.0  # Assign a default value of 0.0 for empty strings

                    if row_value is not None:
                        data.append(float_value)
                    else:
                        break                

                if sorting_order == "ascending":

                    sorted_data = sort.number_sort(data, True)
                
                    data_order = []

                    for i in range(0, len(sorted_data)):
                        temp = sorted_data[i]
                        temp = int(temp)

                        active_sheet = wb['Customers']

                        for i in active_sheet.iter_rows(min_row=2, max_row=active_sheet.max_row, min_col=5, max_col=5):
                            for cell in i:
                                if cell.value == temp and cell.row not in data_order:
                                    row = cell.row
                                    data_order.append(row)

                    queue = Queue()

                    for i in range(0, len(data_order)):
                        queue.enqueue(data_order[i])

                    queue.serialize('queue_state.pkl')

                    window.close()
                    customers_window(queue=True)

                if sorting_order == "descending":

                    sorted_data = sort.number_sort(data, False)
                
                    data_order = []

                    for i in range(0, len(sorted_data)):
                        temp = sorted_data[i]
                        temp = float(temp)

                        active_sheet = wb['Customers']

                        for i in active_sheet.iter_rows(min_row=2, max_row=active_sheet.max_row, min_col=5, max_col=5):
                            for cell in i:
                                if cell.value == temp and cell.row not in data_order:
                                    row = cell.row
                                    data_order.append(row)

                    queue = Queue()

                    for i in range(0, len(data_order)):
                        queue.enqueue(data_order[i])

                    queue.serialize('queue_state.pkl')

                    window.close()
                    customers_window(queue=True)

            if selected_criteria == "Name":

                for row in ws.iter_rows(min_row=2, values_only=True):
                    row_value = row[0]

                    if row_value is not None and not isinstance(row_value, str):
                        row_value = str(row_value)
                    data.append(row_value)

                if sorting_order == "ascending":

                    sorted_data = sort.alphabetical_sort(data, True)
                
                    data_order = []

                    for i in range(0, len(sorted_data)):
                        temp = sorted_data[i]

                        active_sheet = wb['Customers']

                        for i in active_sheet.iter_rows(min_row=2, max_row=active_sheet.max_row, min_col=1, max_col=1):
                            for cell in i:
                                if cell.value == temp and cell.row not in data_order:
                                    row = cell.row
                                    data_order.append(row)

                    queue = Queue()

                    for i in range(0, len(data_order)):
                        queue.enqueue(data_order[i])

                    queue.serialize('queue_state.pkl')

                    window.close()
                    customers_window(queue=True)

                if sorting_order == "descending":

                    sorted_data = sort.alphabetical_sort(data, False)
            
                    data_order = []

                    for i in range(0, len(sorted_data)):
                        temp = sorted_data[i]

                        active_sheet = wb['Customers']

                        for i in active_sheet.iter_rows(min_row=2, max_row=active_sheet.max_row, min_col=1, max_col=1):
                            for cell in i:
                                if cell.value == temp and cell.row not in data_order:
                                    row = cell.row
                                    data_order.append(row)

                    queue = Queue()

                    for i in range(0, len(data_order)):
                        queue.enqueue(data_order[i])

                    queue.serialize('queue_state.pkl')

                    window.close()
                    customers_window(queue=True)
            
    window.close()

def sort_customer_orders_window(customer):
    button_states = {
        "Order Value": 0,
        "Order Quantity": 0
    }

    layout = [
        [sg.Text("Sort By",  pad=15, expand_x=True, font=('Arial Bold', 24), justification='center')],
        [sg.Button("Order Value", pad=0, button_color=('#041c11', '#02b8fa'), mouseover_colors=('#02b8fa', '#041c11'), size=(15, 2), font=('Times New Roman', 15), key="Order Value")],
        [sg.Button("Order Quantity", pad=15, button_color=('#041c11', '#02b8fa'), mouseover_colors=('#02b8fa', '#041c11'), size=(15, 2), font=('Times New Roman', 15), key="Order Quantity")],
        [sg.Text("Least First", font=('Times New Roman', 12)), sg.Slider(range=(0, 1), orientation='h', default_value=0, resolution=1, disable_number_display=True, pad=10, key='-SLIDER-', size=(18, 10), background_color='#041c11'), sg.Text("Most First", font=('Times New Roman', 12))],
        [sg.Button("Back", pad=30, button_color='#041c11', mouseover_colors='#02b8fa', size=(15, 1)), sg.Button("Confirm", pad=30, button_color='#041c11', mouseover_colors='#02b8fa', size=(15, 1))],
        [sg.Exit(button_color='#041c11', size=(15, 1))]
    ]

    window = sg.Window("Sort Customer's Orders Page", layout, size=(500, 450), element_justification='c')

    while True:
        event, values = window.read()

        if event in (sg.WINDOW_CLOSED, "Exit"):
            break

        if event == "Back":
            window.close()
            view_customer_window(customer)

        # Toggle button color
        if event in button_states:
            button_name = event
            
            # Reset all buttons to 0 except the one clicked
            for key in button_states:
                if key != button_name:
                    button_states[key] = 0

            # Toggle the selected button
            button_states[button_name] = 1

            # Update button colors based on states
            for key, state in button_states.items():
                new_color = ('#02b8fa', '#041c11') if state == 1 else ('#041c11', '#02b8fa')
                window[key].update(button_color=new_color)

        if event == "Confirm":
            selected_criteria = [key for key, value in button_states.items() if value == 1][0]
            sorting_order = "ascending" if values['-SLIDER-'] == 0 else "descending"

            sort = sorting_functions()

            wb = openpyxl.load_workbook('Database.xlsx')
            ws = wb['Orders']

            data = []

            if selected_criteria == "Order Value":

                for row in ws.iter_rows(min_row=2, values_only=True):
                    row_value = row[2]

                    try:
                        float_value = float(row_value)
                    except ValueError:
                        float_value = 0.0  # Assign a default value of 0.0 for empty strings

                    if row_value is not None:
                        data.append(float_value)
                    else:
                        break

                if sorting_order == "ascending":

                    sorted_data = sort.number_sort(data, True)
                
                    data_order = []

                    for i in range(0, len(sorted_data)):
                        temp = sorted_data[i]

                        active_sheet = wb['Orders']

                        for i in active_sheet.iter_rows(min_row=2, max_row=active_sheet.max_row, min_col=3, max_col=3):
                            for cell in i:
                                if cell.value == temp and cell.row not in data_order:
                                    row = cell.row
                                    data_order.append(row)

                    queue = Queue()

                    for i in range(0, len(data_order)):
                        queue.enqueue(data_order[i])

                    queue.serialize('queue_state.pkl')

                if sorting_order == "descending":

                    sorted_data = sort.number_sort(data, False)
            
                    data_order = []

                    for i in range(0, len(sorted_data)):
                        temp = sorted_data[i]
                        temp = float(temp)

                        active_sheet = wb['Orders']

                        for i in active_sheet.iter_rows(min_row=2, max_row=active_sheet.max_row, min_col=3, max_col=3):
                            for cell in i:
                                if cell.value == temp and cell.row not in data_order:
                                    row = cell.row
                                    data_order.append(row)

                    queue = Queue()

                    for i in range(0, len(data_order)):
                        queue.enqueue(data_order[i])

                    queue.serialize('queue_state.pkl')

            if selected_criteria == "Order Quantity":

                for row in ws.iter_rows(min_row=2, values_only=True):
                    row_value = row[5]

                    try:
                        int_value = int(row_value)
                    except ValueError:
                        int_value = 0  # Assign a default value of 0 for empty strings

                    if row_value is not None:
                        data.append(int_value)
                    else:
                        break                

                if sorting_order == "ascending":

                    sorted_data = sort.number_sort(data, True)
                
                    data_order = []

                    for i in range(0, len(sorted_data)):
                        temp = sorted_data[i]

                        active_sheet = wb['Orders']

                        for i in active_sheet.iter_rows(min_row=2, max_row=active_sheet.max_row, min_col=6, max_col=6):
                            for cell in i:
                                if cell.value == temp and cell.row not in data_order:
                                    row = cell.row
                                    data_order.append(row)

                    queue = Queue()

                    for i in range(0, len(data_order)):
                        queue.enqueue(data_order[i])

                    queue.serialize('queue_state.pkl')

                if sorting_order == "descending":

                    sorted_data = sort.number_sort(data, False)
                    
                    data_order = []

                    for i in range(0, len(sorted_data)):
                        temp = sorted_data[i]
                        temp = int(temp)

                        active_sheet = wb['Orders']

                        for i in active_sheet.iter_rows(min_row=2, max_row=active_sheet.max_row, min_col=6, max_col=6):
                            for cell in i:
                                if cell.value == temp and cell.row not in data_order:
                                    row = cell.row
                                    data_order.append(row)

                    queue = Queue()

                    for i in range(0, len(data_order)):
                        queue.enqueue(data_order[i])

                    queue.serialize('queue_state.pkl')
            
    window.close()

def sort_vendors_window():
    button_states = {
        "Pending Order Quantity": 0,
        "Total Order Quantity": 0,
        "Name": 0
    }

    layout = [
        [sg.Text("Sort By",  pad=15, expand_x=True, font=('Arial Bold', 24), justification='center')],
        [sg.Button("Pending Order Quantity", pad=0, button_color=('#041c11', '#02b8fa'), mouseover_colors=('#02b8fa', '#041c11'), size=(17, 2), font=('Times New Roman', 15), key="Pending Order Quantity")],
        [sg.Button("Total Order Quantity", pad=15, button_color=('#041c11', '#02b8fa'), mouseover_colors=('#02b8fa', '#041c11'), size=(15, 2), font=('Times New Roman', 15), key="Total Order Quantity")],
        [sg.Button("Name", pad=0, button_color=('#041c11', '#02b8fa'), mouseover_colors=('#02b8fa', '#041c11'), size=(15, 2), font=('Times New Roman', 15), key="Name")],
        [sg.Text("Least/A First", font=('Times New Roman', 12)), sg.Slider(range=(0, 1), orientation='h', default_value=0, resolution=1, disable_number_display=True, pad=10, key='-SLIDER-', size=(18, 10), background_color='#041c11'), sg.Text("Most/Z First", font=('Times New Roman', 12))],
        [sg.Button("Back", pad=30, button_color='#041c11', mouseover_colors='#02b8fa', size=(15, 1)), sg.Button("Confirm", pad=30, button_color='#041c11', mouseover_colors='#02b8fa', size=(15, 1))],
        [sg.Exit(button_color='#041c11', size=(15, 1))]
    ]

    window = sg.Window("Sort Vendors Page", layout, size=(500, 450), element_justification='c')

    while True:
        event, values = window.read()

        if event in (sg.WINDOW_CLOSED, "Exit"):
            break

        if event == "Back":
            window.close()
            vendors_window(queue=False)

        # Toggle button color
        if event in button_states:
            button_name = event
            
            # Reset all buttons to 0 except the one clicked
            for key in button_states:
                if key != button_name:
                    button_states[key] = 0

            # Toggle the selected button
            button_states[button_name] = 1

            # Update button colors based on states
            for key, state in button_states.items():
                new_color = ('#02b8fa', '#041c11') if state == 1 else ('#041c11', '#02b8fa')
                window[key].update(button_color=new_color)

        if event == "Confirm":
            selected_criteria = [key for key, value in button_states.items() if value == 1][0]
            sorting_order = "ascending" if values['-SLIDER-'] == 0 else "descending"

            sort = sorting_functions()

            wb = openpyxl.load_workbook('Database.xlsx')
            ws = wb['Vendors']

            data = []

            if selected_criteria == "Pending Order Quantity":

                for row in ws.iter_rows(min_row=2, values_only=True):
                    row_value = row[6]
                    row_arr = row_value.split(',')

                    n = len(row_arr)

                    try:
                        int_value = int(n)
                    except ValueError:
                        int_value = 0  # Assign a default value of 0 for empty strings

                    if row_value is not None:
                        data.append(int_value)
                    else:
                        break

                if sorting_order == "ascending":

                    sorted_data = sort.number_sort(data, True)
                    
                    data_order = []

                    for i in range(0, len(sorted_data)):
                        temp = sorted_data[i]

                        active_sheet = wb['Vendors']

                        for i in active_sheet.iter_rows(min_row=2, max_row=active_sheet.max_row, min_col=7, max_col=7):
                            for cell in i:
                                comparison = cell.value.split(',')
                                compare = len(comparison)
                                if compare == temp and cell.row not in data_order:
                                    row = cell.row
                                    data_order.append(row)

                    queue = Queue()

                    for i in range(0, len(data_order)):
                        queue.enqueue(data_order[i])

                    queue.serialize('queue_state.pkl')

                    window.close()
                    vendors_window(queue=True)

                if sorting_order == "descending":

                    sorted_data = sort.number_sort(data, False)
                    
                    data_order = []

                    for i in range(0, len(sorted_data)):
                        temp = sorted_data[i]
                        temp = int(temp)

                        active_sheet = wb['Vendors']

                        for i in active_sheet.iter_rows(min_row=2, max_row=active_sheet.max_row, min_col=7, max_col=7):
                            for cell in i:
                                comparison = cell.value.split(',')
                                compare = len(comparison)
                                if compare == temp and cell.row not in data_order:
                                    row = cell.row
                                    data_order.append(row)
                    
                    queue = Queue()

                    for i in range(0, len(data_order)):
                        queue.enqueue(data_order[i])

                    queue.serialize('queue_state.pkl')

                    window.close()
                    vendors_window(queue=True)

            if selected_criteria == "Total Order Quantity":

                for row in ws.iter_rows(min_row=2, values_only=True):
                    curr_row_value = row[6] #Pending Orders
                    past_row_value = row[5] #Completed Orders

                    curr_row_arr = curr_row_value.split(',')
                    past_row_arr = past_row_value.split(',')

                    curr_len = len(curr_row_arr)
                    past_len = len(past_row_arr)

                    n = curr_len + past_len

                    try:
                        int_value = int(n)
                    except ValueError:
                        int_value = 0  # Assign a default value of 0 for empty strings

                    if curr_row_value is not None or past_row_value is not None:
                        data.append(int_value)
                    else:
                        break

                if sorting_order == "ascending":

                    sorted_data = sort.number_sort(data, True)
                    
                    data_order = []

                    for i in range(0, len(sorted_data)):
                        temp = sorted_data[i]

                        active_sheet = wb['Vendors']

                        for row in active_sheet.iter_rows(min_row=2, max_row=active_sheet.max_row, min_col=6, max_col=7):
                            comparison_current = row[1].value
                            comparison_current = comparison_current.split(',')
                            comparison_past = row[0].value  # Fix the index here
                            comparison_past = comparison_past.split(',')
                            compare = len(comparison_past) + len(comparison_current)
                            if compare == temp and row[0].row not in data_order:
                                row = row[0].row
                                data_order.append(row)

                    queue = Queue()

                    for i in range(0, len(data_order)):
                        queue.enqueue(data_order[i])

                    queue.serialize('queue_state.pkl')

                    window.close()
                    vendors_window(queue=True)

                if sorting_order == "descending":

                    sorted_data = sort.number_sort(data, False)
                    
                    data_order = []

                    for i in range(0, len(sorted_data)):
                        temp = sorted_data[i]
                        temp = int(temp)

                        active_sheet = wb['Vendors']

                        for row in active_sheet.iter_rows(min_row=2, max_row=active_sheet.max_row, min_col=6, max_col=7):
                            comparison_current = row[1].value
                            comparison_current = comparison_current.split(',')
                            comparison_past = row[0].value  # Fix the index here
                            comparison_past = comparison_past.split(',')
                            compare = len(comparison_past) + len(comparison_current)
                            if compare == temp and row[0].row not in data_order:
                                row = row[0].row
                                data_order.append(row)

                    queue = Queue()

                    for i in range(0, len(data_order)):
                        queue.enqueue(data_order[i])

                    queue.serialize('queue_state.pkl')

                    window.close()
                    vendors_window(queue=True)

            if selected_criteria == "Name":

                for row in ws.iter_rows(min_row=2, values_only=True):
                    row_value = row[0]

                    if row_value is not None and not isinstance(row_value, str):
                        row_value = str(row_value)
                    data.append(row_value)

                if sorting_order == "ascending":

                    sorted_data = sort.alphabetical_sort(data, False)
                    
                    data_order = []

                    for i in range(0, len(sorted_data)):
                        temp = sorted_data[i]

                        active_sheet = wb['Vendors']

                        for i in active_sheet.iter_rows(min_row=2, max_row=active_sheet.max_row, min_col=1, max_col=1):
                            for cell in i:
                                if cell.value == temp and cell.row not in data_order:
                                    row = cell.row
                                    data_order.append(row)

                    queue = Queue()

                    for i in range(0, len(data_order)):
                        queue.enqueue(data_order[i])

                    queue.serialize('queue_state.pkl')

                    window.close()
                    vendors_window(queue=True)

                if sorting_order == "descending":

                    sorted_data = sort.alphabetical_sort(data, True)
                
                    data_order = []

                    for i in range(0, len(sorted_data)):
                        temp = sorted_data[i]

                        active_sheet = wb['Vendors']

                        for i in active_sheet.iter_rows(min_row=2, max_row=active_sheet.max_row, min_col=1, max_col=1):
                            for cell in i:
                                if cell.value == temp and cell.row not in data_order:
                                    row = cell.row
                                    data_order.append(row)

                    queue = Queue()

                    for i in range(0, len(data_order)):
                        queue.enqueue(data_order[i])

                    queue.serialize('queue_state.pkl')

                    window.close()
                    vendors_window(queue=True)
            
    window.close()

def valid(data_string, lenght):
    num_characters = sum([len(i) for i in data_string.split()])

    if num_characters == lenght:
        return True
    else:
        return False

def create_order_window():

    default_date = datetime.now() + timedelta(weeks=3)
    default_date_str = default_date.strftime("%d-%m-%Y")
    
    layout = [
        [sg.Text("New Order",  pad=15, expand_x=True, font=('Arial Bold', 24), justification='center')],
        [sg.Text("Client Phone Number: ", expand_x=True, font=('Times New Roman', 15), justification='center'), sg.InputText(key="-CLIENT-")],
        [sg.Text("Vendor Phone Number:", expand_x=True, font=('Times New Roman', 15), justification='center'), sg.InputText(key="-VENDOR-")],
        [sg.Text("Value: ", expand_x=True, font=('Times New Roman', 15), justification='center'), sg.InputText(key="-VALUE-")],
        [sg.Text("Completion Status: ", expand_x=True, font=('Times New Roman', 15), justification='center'), sg.Text("0% Complete", expand_x=True, font=('Times New Roman', 15), justification='center') ,sg.Slider(range=(0, 100), orientation='h', default_value=0, resolution=10, key='-SLIDER-', size=(18, 10)), sg.Text("100% Complete", expand_x=True, font=('Times New Roman', 15), justification='center')],
        [sg.Text("Order Description: ", expand_x=True, font=('Times New Roman', 15), justification='center'), sg.InputText(key="-DESCRIPTION-")],
        [sg.Text("Quantity: ", expand_x=True, font=('Times New Roman', 15), justification='center'),sg.InputText(key="-QUANTITY-")],
        [sg.Text("Due Date: ", expand_x=True, font=('Times New Roman', 15), justification='center'), sg.InputText(default_text=default_date_str, key="-DUE_DATE-")],
        [sg.CalendarButton("Change Date", key="-CALENDAR-", target="-DUE_DATE-", default_date_m_d_y=(default_date.month, default_date.day, default_date.year), pad=15, button_color='#041c11', format="%d-%m-%Y", size=(15, 1))],
        [sg.Button("Back", pad=15, button_color='#041c11', mouseover_colors='#02b8fa', size=(15, 1)), sg.Button("Confirm", pad=15, button_color='#041c11', mouseover_colors='#02b8fa', size=(15, 1))],
        [sg.Exit(button_color='#041c11', size=(15, 1))]
    ]

    window = sg.Window("Add New Order Page", layout, size=(625, 450), element_justification='c')

    while True:
        event, values = window.read()
        if event in (sg.WINDOW_CLOSED, "Exit"):
            break
        if event == "Back":
            window.close()
            orders_window(queue=False)

        if event == "Due Date":
            due_date_str = values["-DUE_DATE-"]
            window['Add New Order Page'].update(default_date_str)

        if event == "Confirm":
            client = values["-CLIENT-"]
            vendor = values["-VENDOR-"]
            value = values["-VALUE-"]
            status = values["-SLIDER-"]
            description = values["-DESCRIPTION-"]
            quantity = values["-QUANTITY-"]
            due_date = values["-DUE_DATE-"]

            if valid(client, 10) and valid(vendor, 10):

                # Split the input date string into day, month, and year
                day, month, year = due_date.split("-")
                # Format the date as "ddmmyyyy"
                final_due_date = f"{day}{month}{year}"

                #Create and Store order
                new_order = Order(Client=client, Vendor=vendor, Value=value, Status=status, 
                                  Description=description, Quantity=quantity, due_date=final_due_date)
                
                orderid = new_order.create_order_id()
                new_order.link_client(client, orderid)
                #new_order.link_vendor(vendor, orderid)

                new_order.storeOrder()

                sg.popup("Order Stored Successfully!")
                window.close()
                orders_window(queue=False)

            else:
                sg.popup("Invalid Phone Number entered")

    window.close()

def add_customer_window():
    
    layout = [
        [sg.Text("Add Customer",  pad=15, expand_x=True, font=('Arial Bold', 24), justification='center')],
        [sg.Text("Customer Name: ", pad=15, expand_x=True, font=('Times New Roman', 15), justification='center'), sg.InputText(key="-NAME-")],
        [sg.Text("Customer Phone Number:", pad=15, expand_x=True, font=('Times New Roman', 15), justification='center'), sg.InputText(key="-PHONE-")],
        [sg.Text("Additional Information: ", pad=15, expand_x=True, font=('Times New Roman', 15), justification='center'), sg.InputText(key="-ADDINFO-")],
        [sg.Text("Referral Location: ", pad=15, expand_x=True, font=('Times New Roman', 15), justification='center'), sg.InputText(key="-REFERRAL-")],
        [sg.Button("Back", pad=15, button_color='#041c11', mouseover_colors='#02b8fa', size=(15, 1)), sg.Button("Confirm", pad=15, button_color='#041c11', mouseover_colors='#02b8fa', size=(15, 1))],
        [sg.Exit(button_color='#041c11', size=(15, 1))]
    ]

    window = sg.Window("Add New Customer Page", layout, size=(625, 400), element_justification='c')

    while True:
        event, values = window.read()
        if event in (sg.WINDOW_CLOSED, "Exit"):
            break
        if event == "Back":
            window.close()
            customers_window(queue=False)
        if event == "Confirm":
            name = values["-NAME-"]
            phone_num = values["-PHONE-"]
            additional_info = values["-ADDINFO-"]
            referral = values["-REFERRAL-"]

            if valid(phone_num, 10):

                new_customer = Customer(Name=name, Phone_Num=phone_num, Referral=referral, AdditionalInfo=additional_info)
                new_customer.store_customer()

                sg.popup("Customer Stored Successfully!")

                window.close()
                customers_window(queue=False)

            else:
                sg.popup("Invalid Phone Number Entered")

    window.close()

def add_vendor_window():
    
    layout = [
        [sg.Text("Add Vendor",  pad=15, expand_x=True, font=('Arial Bold', 24), justification='center')],
        [sg.Text("Vendor Name: ", pad=15, expand_x=True, font=('Times New Roman', 15), justification='center'), sg.InputText(key="-NAME-")],
        [sg.Text("Vendor Phone Number:", pad=15, expand_x=True, font=('Times New Roman', 15), justification='center'), sg.InputText(key="-PHONE-")],
        [sg.Text("Additional Information: ", pad=15, expand_x=True, font=('Times New Roman', 15), justification='center'), sg.InputText(key="-ADDINFO-")],
        [sg.Text("Vendor Address: ", pad=15, expand_x=True, font=('Times New Roman', 15), justification='center'), sg.InputText(key="-ADDRESS-")],
        [sg.Button("Back", pad=15, button_color='#041c11', mouseover_colors='#02b8fa', size=(15, 1)), sg.Button("Confirm", pad=15, button_color='#041c11', mouseover_colors='#02b8fa', size=(15, 1))],
        [sg.Exit(button_color='#041c11', size=(15, 1))]
    ]

    window = sg.Window("Add New Vendor Page", layout, size=(600, 400), element_justification='c')

    while True:
        event, values = window.read()
        if event in (sg.WINDOW_CLOSED, "Exit"):
            break
        if event == "Back":
            window.close()
            vendors_window(queue=False)
        if event == "Confirm":
            name = values['-NAME-']
            phone_num = values['-PHONE-']
            additional_info = values['-ADDINFO-']
            address = values['-ADDRESS-']

            if valid(phone_num, 10):

                new_vendor = Vendor(Name=name, PhoneNumber=phone_num, Address=address, AdditionalInfo=additional_info)
                new_vendor.store_vendor()

                sg.popup("Vendor Stored Successfully!")

                window.close()
                vendors_window(queue=False)

            else:
                sg.popup("Invalid Phone Number Entered")

    window.close()

def view_order_window(row_num):
    
    wb = openpyxl.load_workbook('Database.xlsx')
    ws = wb['Orders']

    row_values = []
    for row in ws.iter_rows(min_row=row_num, max_row=row_num, values_only=True):
        row_values.extend(row)

    wb.close()

    customer = row[0]
    vendor = row[1]
    value = f"{row[2]}"
    status = row[3]
    description = row[4]
    quantity = f"{row[5]}"
    due_date = f"{row[7]}"

    if due_date is not None and due_date != "":
        day, month, year = due_date[:2], due_date[2:4], due_date[4:]
        due_date = f"{day}-{month}-{year}"
    
    layout = [
        [sg.Text("View Order", pad=15, expand_x=True, font=('Arial Bold', 22), justification='center'), sg.Text("Editing off", expand_x=True, font=('Times New Roman', 15), justification='center'), sg.Slider(range=(0, 1), default_value=0, resolution=1, orientation='h', disable_number_display=True, pad=10, key='-EDITSLIDER-', size=(18, 10), background_color='#041c11'), sg.Text("Editing on", expand_x=True, font=('Times New Roman', 15), justification='center')],
        [sg.Text("Client Number: ", pad=15, expand_x=True, font=('Times New Roman', 15), justification='center'), sg.InputText(default_text=customer, key="-CLIENT-")],
        [sg.Text("Vendor Number:", pad=15, expand_x=True, font=('Times New Roman', 15), justification='center'), sg.InputText(default_text=vendor, key="-VENDOR-")],
        [sg.Text("Value: ", pad=15, expand_x=True, font=('Times New Roman', 15), justification='center'), sg.InputText(default_text=value, key="-VALUE-")],
        [sg.Text("Status: ", pad=15, expand_x=True, font=('Times New Roman', 15), justification='center'), sg.Text("0% Complete", expand_x=True, font=('Times New Roman', 15), justification='center'), sg.Slider(range=(0, 100), resolution=10, orientation='h', pad=10, key='-STATUSSLIDER-', size=(18, 10), default_value=status), sg.Text("100% Complete", expand_x=True, font=('Times New Roman', 15), justification='center')],
        [sg.Text("Description: ", pad=15, expand_x=True, font=('Times New Roman', 15), justification='center'), sg.InputText(default_text=description, key="-DESCRIPTION-")],
        [sg.Text("Quantity: ", pad=15, expand_x=True, font=('Times New Roman', 15), justification='center'), sg.InputText(default_text=quantity, key="-QUANTITY-")],
        [sg.Text("Due Date: ", expand_x=True, font=('Times New Roman', 15), justification='center'), sg.InputText(default_text=due_date, key="-DUE_DATE-")],
        [sg.CalendarButton("Change Date", key="-CALENDAR-", target="-DUE_DATE-", pad=15, button_color='#041c11', format="%d-%m-%Y", size=(15, 1))],
        [sg.Button("Delete Order", pad=15, button_color='#041c11', mouseover_colors='#02b8fa', size=(15, 1)), sg.Button("Back", pad=15, button_color='#041c11', mouseover_colors='#02b8fa', size=(15, 1)), sg.Button("Save Changes", pad=15, button_color='#041c11', mouseover_colors='#02b8fa', size=(15, 1))],
        [sg.Exit(button_color='#041c11', size=(15, 1))]
    ]

    window = sg.Window("View Order Page", layout, size=(600, 600), element_justification='c')

    while True:
        event, values = window.read()
        if event in (sg.WINDOW_CLOSED, "Exit"):
            break
        if event == "Back":
            window.close()
            orders_window(queue=False)
        if event == "Delete Order":
            window.close()
            delete_order_page(row_num)
        if event == "Save Changes" and values["-EDITSLIDER-"] == 1:

            client = values["-CLIENT-"]
            vendor = values["-VENDOR-"]
            value = float(values["-VALUE-"])
            status = int(values["-STATUSSLIDER-"])
            description = values["-DESCRIPTION-"]
            quantity = int(values["-QUANTITY-"])
            due_date = values["-DUE_DATE-"]

            if valid(client, 10) and valid(vendor, 10):

                # Split the input date string into day, month, and year
                day, month, year = due_date.split("-")
                # Format the date as "ddmmyyyy"
                final_due_date = f"{day}{month}{year}"
                final_due_date = int(final_due_date)
                
                wb = openpyxl.load_workbook('Database.xlsx')
                ws = wb['Orders']

                ws.cell(row=row_num, column=1, value=client)
                ws.cell(row=row_num, column=2, value=vendor)
                ws.cell(row=row_num, column=3, value=value)
                ws.cell(row=row_num, column=4, value=status)
                ws.cell(row=row_num, column=5, value=description)
                ws.cell(row=row_num, column=6, value=quantity)
                ws.cell(row=row_num, column=8, value=final_due_date)
                wb.save('Database.xlsx')

                sg.popup("Changes Saved!")

            else:
                sg.popup("Invalid Changes Made")

    window.close()

def view_customer_window(row_num):

    wb = openpyxl.load_workbook('Database.xlsx')
    ws = wb['Customers']

    row_values = []
    for row in ws.iter_rows(min_row=row_num, max_row=row_num, values_only=True):
        row_values.extend(row)

    name = row[0]
    phone_num = row[1]
    referral = row[2]
    additional_info = row[5]

    if row[6] is not None:
        past_orders = row[6].split(',')
    else:
        past_orders = []

    customer = Customer(name, phone_num, referral, additional_info)

    orders_placed = f"{customer.calc_orders_placed(past_orders)}"
    total_spending = f"{customer.calc_total_spending(past_orders)}"
    
    layout = [
        [sg.Text("View Customer", pad=15, expand_x=True, font=('Arial Bold', 22), justification='center'), sg.Text("Editing off", expand_x=True, font=('Times New Roman', 15), justification='center'), sg.Slider(range=(0, 1), default_value=0, resolution=1, orientation='h', disable_number_display=True, pad=10, key='-EDITSLIDER-', size=(18, 10), background_color='#041c11'), sg.Text("Editing on", expand_x=True, font=('Times New Roman', 15), justification='center')],
        [sg.Text("Customer Name: ", pad=15, expand_x=True, font=('Times New Roman', 15), justification='center'), sg.InputText(default_text=name, key="-CLIENT-"), sg.Button("Past Orders", button_color='#041c11', mouseover_colors='#02b8fa', size=(15, 1))],
        [sg.Text("Customer Phone Number:", pad=15, expand_x=True, font=('Times New Roman', 15), justification='center'), sg.InputText(default_text=phone_num, key="-PHONE-")],
        [sg.Text("Referral Location: ", pad=15, expand_x=True, font=('Times New Roman', 15), justification='center'), sg.InputText(default_text=referral, key="-REFERRAL-")],
        [sg.Text(f"Total Orders Placed: {orders_placed}", pad=15, expand_x=True, font=('Times New Roman', 15), justification='center')],
        [sg.Text(f"Total Spending: {total_spending}", pad=15, expand_x=True, font=('Times New Roman', 15), justification='center')],
        [sg.Text("Additional Information", pad=15, expand_x=True, font=('Times New Roman', 15), justification='center'), sg.InputText(default_text=additional_info, key="-ADDINFO-")],
        [sg.Button("Back", pad=15, button_color='#041c11', mouseover_colors='#02b8fa', size=(15, 1)), sg.Button("Save Changes", pad=15, button_color='#041c11', mouseover_colors='#02b8fa', size=(15, 1))],
        [sg.Exit(button_color='#041c11', size=(15, 1))]
    ]

    window = sg.Window("View Customer Page", layout, size=(650, 500), element_justification='c')

    while True:
        event, values = window.read()
        if event in (sg.WINDOW_CLOSED, "Exit"):
            break
        if event == "Back":
            window.close()
            customers_window(queue=False)
        if event == "Save Changes" and values["-EDITSLIDER-"] == 1:
            
            name = values["-CLIENT-"]
            phone_num = values["-PHONE-"]
            referral = values["-REFERRAL-"]
            additional_info = values["-ADDINFO-"]

            if valid(phone_num, 10):

                wb = openpyxl.load_workbook('Database.xlsx')
                ws = wb['Customers']

                ws.cell(row=row_num, column=1, value=name)
                ws.cell(row=row_num, column=2, value=phone_num)
                ws.cell(row=row_num, column=3, value=referral)
                ws.cell(row=row_num, column=6, value=additional_info)

                wb.save('Database.xlsx')

                sg.popup("Changes Saved!")

            else:
                sg.popup("Invalid Data Entered")

        if event == "Past Orders":
            window.close()
            customer_orders_window(row_num, queue=False)

    window.close()

def view_vendor_window(row_num):

    wb = openpyxl.load_workbook('Database.xlsx')
    ws = wb['Vendors']

    row_values = []
    for row in ws.iter_rows(min_row=row_num, max_row=row_num, values_only=True):
        row_values.extend(row)

    name = row[0]
    phone_num = row[1]
    address = row[2]
    status = row[3]
    additional_info = row[4]
    
    layout = [
        [sg.Text("View Vendor",  pad=15, expand_x=True, font=('Arial Bold', 22), justification='center'), sg.Text("Editing off", expand_x=True, font=('Times New Roman', 15), justification='center'), sg.Slider(range=(0, 1), default_value=0, resolution=1, orientation='h', disable_number_display=True, pad=10, key='-EDITSLIDER-', size=(18, 10), background_color='#041c11'), sg.Text("Editing on", expand_x=True, font=('Times New Roman', 15), justification='center')],
        [sg.Text("Vendor Name: ", pad=15, expand_x=True, font=('Times New Roman', 15), justification='center'), sg.InputText(default_text=name, key="-VENDOR-"), sg.Text(f"Vendor Status: {status}", pad=15, expand_x=True, font=('Times New Roman', 15), justification='center')],
        [sg.Text("Vendor Phone Number:", pad=15, expand_x=True, font=('Times New Roman', 15), justification='center'), sg.InputText(default_text=phone_num, key="-PHONE-"), sg.Button("Vendor Orders", pad=15, button_color='#041c11', mouseover_colors='#02b8fa', size=(15, 1))],
        [sg.Text("Additional Information: ", pad=15, expand_x=True, font=('Times New Roman', 15), justification='center'), sg.InputText(default_text=additional_info, key="-INFO-")],
        [sg.Text("Address: ", pad=15, expand_x=True, font=('Times New Roman', 15), justification='center'), sg.InputText(default_text=address, key="-ADDRESS-")],
        [sg.Button("Back", pad=15, button_color='#041c11', mouseover_colors='#02b8fa', size=(15, 1)), sg.Button("Save Changes", pad=15, button_color='#041c11', mouseover_colors='#02b8fa', size=(15, 1))],
        [sg.Exit(button_color='#041c11', size=(15, 1))]
    ]

    window = sg.Window("View Vendor Page", layout, size=(700, 400), element_justification='c')

    while True:
        event, values = window.read()
        if event in (sg.WINDOW_CLOSED, "Exit"):
            break
        if event == "Back":
            window.close()
            vendors_window(queue=False)
        if event == "Save Changes" and values["-EDITSLIDER-"] == 1:
            
            name = values["-VENDOR-"]
            phone_num = values["-PHONE-"]
            address = values["-ADDRESS-"]
            additional_info = values["-INFO-"]

            if valid(phone_num, 10):

                wb = openpyxl.load_workbook('Database.xlsx')
                ws = wb['Vendors']

                ws.cell(row=row_num, column=1, value=name)
                ws.cell(row=row_num, column=2, value=phone_num)
                ws.cell(row=row_num, column=3, value=address)
                ws.cell(row=row_num, column=5, value=additional_info)

                wb.save('Database.xlsx')

                sg.popup("Changes Saved!")

            else:
                sg.popup("Invalid Data Entered")

        if event == "Vendor Orders":
            window.close()
            vendor_orders_window(row_num)

    window.close()

def customer_orders_window(customer, queue):
    
    wb = openpyxl.load_workbook('Database.xlsx')
    customers_sheet = wb['Customers']
    orders_sheet = wb['Orders']

    past_orders = customers_sheet.cell(row=customer, column=7).value

    try:
        past_orders = list(past_orders.split(','))
    except AttributeError:
        past_orders = []

    text_elements = []
    button_elements = []

    if queue == False:
        for past_order in range(0, len(past_orders)):
            order_text = sg.Text(f"Order {past_order + 1}", pad=5)
            button = sg.Button("View/Edit Order", key=f"BUTTON{past_order + 1}")
            text_elements.append(order_text)
            button_elements.append(button)

    if queue == True:
        restored = Queue.deserialize('queue_state.pkl')

        lst = []

        while restored.is_empty() == False:
            temp = restored.dequeue()
            lst.append(temp)

        for i in range(0, len(lst)):
            order_text = sg.Text(f"Order {lst[i]}", pad=5)
            button = sg.Button("View/Edit Order", key=f"BUTTON{lst[i] - 1}")
            text_elements.append(order_text)
            button_elements.append(button)

    col_layout = list(zip(text_elements, button_elements))

    layout = [
        [sg.Button("Back To Customer", pad=15, button_color='#041c11', size=(7, 3), mouseover_colors='#02b8fa'), sg.Text("Past Orders", pad=15, expand_x=True, font=('Arial Bold', 24), justification='center'), sg.Exit(pad=15, button_color='#041c11', size=(7, 3))],
        [sg.Button("Sort Past Orders", pad=10, button_color='#041c11', mouseover_colors='#02b8fa', size=(15, 1))],
        [sg.Column(col_layout, scrollable=True, size=(500, 275), vertical_scroll_only=True, element_justification='center')],
    ]

    window = sg.Window("View Customer Orders Page", layout, size=(650, 400), element_justification='c')

    while True:
        event, values = window.read()
        if event in (sg.WINDOW_CLOSED, "Exit"):
            break
        if event == "Back To Customer":
            window.close()
            view_customer_window(customer)
        if event == "Sort Past Orders":
            window.close()
            sort_customer_orders_window(customer=customer)
        if event.startswith("BUTTON"):
            order_num = int(event.replace("BUTTON", "")) - 1
            order_row = find_entry_info('Orders', past_orders[order_num], 7)
            window.close()
            view_order_window(order_row)

    window.close()

def vendor_orders_window(vendor):
    
    wb = openpyxl.load_workbook('Database.xlsx')
    vendors_sheet = wb['Vendors']
    orders_sheet = wb['Orders']

    current_orders = vendors_sheet.cell(row=vendor, column=7).value
    current_orders = list(current_orders.split(','))

    text_elements = []
    button_elements = []

    for order in range(0, len(current_orders)):
        order_text = sg.Text(f"Order {order + 1}", pad=5)
        button = sg.Button("View/Edit Order", key=f"BUTTON{order + 1}")
        text_elements.append(order_text)
        button_elements.append(button)

    col_layout = list(zip(text_elements, button_elements))

    layout = [
        [sg.Button("Back To Vendor", pad=15, button_color='#041c11', size=(7, 3), mouseover_colors='#02b8fa'), sg.Text("Vendor Orders",  pad=15, expand_x=True, font=('Arial Bold', 24), justification='center'), sg.Exit(pad=15, button_color='#041c11', size=(7, 3))],
        [sg.Column(col_layout, scrollable=True, size=(500, 275), vertical_scroll_only=True, element_justification='center')]
    ]

    window = sg.Window("View Vendors Orders Page", layout, size=(650, 400), element_justification='c')

    while True:
        event, values = window.read()
        if event in (sg.WINDOW_CLOSED, "Exit"):
            break
        if event == "Back To Vendor":
            window.close()
            view_vendor_window(vendor)
        if event.startswith("BUTTON"):
            order_num = int(event.replace("BUTTON", "")) - 1
            order_row = find_entry_info('Orders', current_orders[order_num], 7)
            window.close()
            view_order_window(order_row)

    window.close()

def delete_order_page(row_num):

    layout = [
        [sg.Text("Delete Order",  pad=15, expand_x=True, font=('Arial Bold', 22), justification='center')],
        [sg.Text("Are you sure you want to delete this order?", pad=15, expand_x=True, font=('Times New Roman', 15), justification='center')],
        [sg.Text("This action cannot be undone", pad=15, expand_x=True, font=('Times New Roman', 15), justification='center')],
        [sg.Button("Back", pad=15, button_color='#041c11', mouseover_colors='#02b8fa', size=(15, 1)), sg.Button("Confirm", pad=15, button_color='#041c11', mouseover_colors='#02b8fa', size=(15, 1))],
        [sg.Exit(button_color='#041c11', size=(15, 1))]
    ]

    window = sg.Window("Delete Order Page", layout, size=(600, 350), element_justification='c')

    while True:
        event, values = window.read()
        if event in (sg.WINDOW_CLOSED, "Exit"):
            break
        if event == "Back":
            window.close()
            view_order_window(row_num)
        if event == "Confirm":
            delete_entry(row_num)
            sg.popup("Order Successfully Deleted")
            window.close()
            orders_window(queue=False)

    window.close()

#main_window()