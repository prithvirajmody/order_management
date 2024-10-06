from GUI import *
import threading
import time
import openpyxl
from datetime import datetime, timedelta
from Notification import send_message


def is_within_one_week(date1_str, date2_str):
    # Convert date strings to datetime objects
    date_format = "%d%m%Y"
    date1 = datetime.strptime(date1_str, date_format)
    date2 = datetime.strptime(date2_str, date_format)

    # Calculate the difference between the two dates
    time_difference = abs(date2 - date1)

    # Check if the time difference is less than a week (7 days)
    one_week = timedelta(days=7)
    return time_difference < one_week

def notification_thread():

    current_date = datetime.now()
    current_date = current_date.strftime("%d%m%Y")

    wb = openpyxl.load_workbook('Database.xlsx')
    ws = wb['Orders']

    for row in ws.iter_rows(2, max_row=ws.max_row):
        for cell in row:
            status = row[3].value
            due_date = row[7].value
            due_date = str(due_date)

        if status <= 50 and is_within_one_week(current_date, due_date) == True:
            #print(f"Order Number {cell.row - 1}")
            send_message(message_body=f"Order Number {cell.row - 1} needs Attention!")

    time.sleep(86400)   #Thread runs once a day

background_thread = threading.Thread(target=notification_thread)
#background_thread.start()

if __name__ == "__main__":
    main_window()